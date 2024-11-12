import os
import json
import time
import csv
import sqlite3
import schedule
import asyncio
import openpyxl
from datetime import datetime, timedelta
from telethon import TelegramClient
from telethon.tl.functions.stories import GetAllStoriesRequest
from telethon.tl.types import MessageMediaPhoto, MessageMediaDocument
from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.prompt import Prompt, IntPrompt
from rich.table import Table
from rich import box
from rich.align import Align

console = Console()

def display_banner():
    """Display the application banner"""
    console.print(Panel("""[bold magenta]
████████╗ ██████╗ ███████╗███████╗ 
╚══██╔══╝██╔════╝ ██╔════╝██╔════╝ 
   ██║   ██║  ███╗███████╗███████╗ 
   ██║   ██║   ██║╚════██║╚════██║ 
   ██║   ╚██████╔╝███████║███████║ 
   ╚═╝    ╚═════╝ ╚══════╝╚══════╝ 
[/bold magenta]
[bold cyan]═══════════════════════════════════════════════════════════════════════════════════════
                        Telegram Story Scraper
═══════════════════════════════════════════════════════════════════════════════════════[/bold cyan]""",
    box=box.DOUBLE_EDGE,
    border_style="bright_blue"))

def display_export_banner():
    """Display the export menu banner"""
    console.print(Panel("""[bold cyan]
╔═══════════════════════════════════════════════════════════════════════════╗
║                            EXPORT DATA MENU                               ║
╚═══════════════════════════════════════════════════════════════════════════╝[/bold cyan]""",
    box=box.DOUBLE_EDGE,
    border_style="bright_blue"))

class StoryScraper:
    def __init__(self):
        self.db_file = 'stories.db'
        self.excel_file_path = 'stories_info.xlsx'
        self.csv_file_path = 'stories_info.csv'
        self.credentials_file = 'credentials.json'
        self.credentials = None
        self.client = None
        self.console = Console()
        self.initialize_database()

    async def initialize_client(self):
        """Initialize and authenticate the Telegram client"""
        try:
            if not self.client:
                console.print("[cyan]Initializing Telegram client...[/cyan]")
                self.credentials = self.load_credentials()
                
                self.client = TelegramClient('session_name', 
                                           self.credentials['api_id'], 
                                           self.credentials['api_hash'])
                
                await self.client.connect()
                
                if not await self.client.is_user_authorized():
                    console.print("[yellow]Requesting authentication code...[/yellow]")
                    await self.client.send_code_request(self.credentials['phone_number'])
                    code = Prompt.ask("[cyan]Enter the code you received on Telegram")
                    try:
                        await self.client.sign_in(self.credentials['phone_number'], code)
                        console.print("[green]Successfully authenticated![/green]")
                    except Exception as e:
                        console.print(f"[red]Failed to sign in: {str(e)}[/red]")
                        return False
                
                console.print("[green]✓[/green] Client initialized and authenticated!")
                return True
                
        except Exception as e:
            console.print(f"[red]Error during client initialization: {str(e)}[/red]")
            return False

    def load_credentials(self):
        """Load or prompt for credentials"""
        if os.path.exists(self.credentials_file):
            with open(self.credentials_file, 'r') as f:
                return json.load(f)
        else:
            return self.prompt_for_credentials()

    def prompt_for_credentials(self):
        """Prompt user for Telegram API credentials"""
        console.print(Panel("[yellow]Please enter your Telegram API credentials[/yellow]"))
        credentials = {
            'api_id': Prompt.ask("Enter your API ID"),
            'api_hash': Prompt.ask("Enter your API Hash"),
            'phone_number': Prompt.ask("Enter your phone number (with country code)")
        }
        
        with open(self.credentials_file, 'w') as f:
            json.dump(credentials, f)
        
        return credentials

    def initialize_database(self):
        """Initialize SQLite database"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS stories (
            user_id INTEGER,
            story_id INTEGER PRIMARY KEY,
            timestamp TEXT,
            filename TEXT
        )
        ''')
        conn.commit()
        conn.close()

    def fetch_stories_from_db(self):
        """Fetch existing stories from database"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        cursor.execute('SELECT user_id, story_id FROM stories')
        stories = cursor.fetchall()
        conn.close()
        return set(stories)

    def insert_story(self, user_id, story_id, timestamp, filename):
        """Insert a new story into database"""
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        cursor.execute('''
        INSERT OR IGNORE INTO stories (user_id, story_id, timestamp, filename)
        VALUES (?, ?, ?, ?)
        ''', (user_id, story_id, timestamp, filename))
        conn.commit()
        conn.close()
    async def scrape_stories(self):
        """Scrape stories from Telegram"""
        if not self.client:
            success = await self.initialize_client()
            if not success:
                return
        
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console
        ) as progress:
            main_task = progress.add_task("[cyan]Scanning for stories...", total=None)
            
            try:
                all_stories = await self.client(GetAllStoriesRequest())

                if not all_stories or not all_stories.peer_stories:
                    console.print("[yellow]No stories found.[/yellow]")
                    return

                existing_stories = self.fetch_stories_from_db()
                new_stories_count = 0
                total_stories = sum(len(peer_story.stories) for peer_story in all_stories.peer_stories)
                
                download_task = progress.add_task("[cyan]Downloading media...", total=total_stories)

                for peer_story in all_stories.peer_stories:
                    user_id = peer_story.peer.user_id
                    
                    for story in peer_story.stories:
                        story_id = story.id
                        if (user_id, story_id) in existing_stories:
                            progress.advance(download_task)
                            continue

                        timestamp_utc = story.date
                        timestamp_local = timestamp_utc + timedelta(hours=2)
                        timestamp = timestamp_local.strftime('%Y-%m-%d %H:%M:%S')
                        
                        try:
                            media = story.media
                            filename = None

                            if isinstance(media, MessageMediaPhoto):
                                filename = f"stories/{user_id}_{story_id}.jpg"
                                await self.client.download_media(media.photo, file=filename)
                            elif isinstance(media, MessageMediaDocument):
                                ext = media.document.mime_type.split('/')[1]
                                filename = f"stories/{user_id}_{story_id}.{ext}"
                                await self.client.download_media(media.document, file=filename)

                            if filename:
                                self.insert_story(user_id, story_id, timestamp, filename)
                                new_stories_count += 1
                                console.print(f"[green]Downloaded:[/green] {filename}")
                                
                        except Exception as e:
                            console.print(f"[red]Error downloading story {story_id}: {str(e)}[/red]")

                        progress.advance(download_task)

                progress.update(main_task, completed=True)
                if new_stories_count > 0:
                    console.print(f"[green]✓[/green] Downloaded {new_stories_count} new stories!")
                else:
                    console.print("[yellow]No new stories to download[/yellow]")

            except Exception as e:
                console.print(f"[red]Error during scraping: {str(e)}[/red]")
                return False

            return True

    def export_to_excel(self):
        """Export stories to Excel file"""
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console
        ) as progress:
            task = progress.add_task("[cyan]Preparing Excel export...", total=None)
            
            try:
                conn = sqlite3.connect(self.db_file)
                cursor = conn.cursor()
                cursor.execute('SELECT * FROM stories ORDER BY timestamp DESC')
                stories = cursor.fetchall()
                conn.close()

                if not stories:
                    console.print("[yellow]No stories to export[/yellow]")
                    return

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Stories"
                
                headers = ["User ID", "Story ID", "Timestamp", "Filename"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)

                for row, story in enumerate(stories, 2):
                    for col, value in enumerate(story, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = value
                        if row % 2:
                            cell.fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                excel_file = f"stories_export_{timestamp}.xlsx"
                wb.save(excel_file)
                progress.update(task, completed=True)
                console.print(f"[green]✓[/green] Successfully exported to {excel_file}!")

            except Exception as e:
                console.print(f"[red]Error during Excel export: {str(e)}[/red]")

    def export_to_csv(self):
        """Export stories to CSV file"""
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console
        ) as progress:
            task = progress.add_task("[cyan]Preparing CSV export...", total=None)
            
            try:
                conn = sqlite3.connect(self.db_file)
                cursor = conn.cursor()
                cursor.execute('SELECT * FROM stories ORDER BY timestamp DESC')
                stories = cursor.fetchall()
                conn.close()

                if not stories:
                    console.print("[yellow]No stories to export[/yellow]")
                    return

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                csv_file = f"stories_export_{timestamp}.csv"

                with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
                    csv_writer = csv.writer(csvfile)
                    csv_writer.writerow(["User ID", "Story ID", "Timestamp", "Filename"])
                    csv_writer.writerows(stories)

                progress.update(task, completed=True)
                console.print(f"[green]✓[/green] Successfully exported to {csv_file}!")

            except Exception as e:
                console.print(f"[red]Error during CSV export: {str(e)}[/red]")

    def export_data(self):
        """Export data menu"""
        while True:
            console.clear()
            display_export_banner()
            
            menu = Table.grid(padding=1)
            menu.add_row("[cyan][[/cyan]1[cyan]][/cyan] [bold]Export to Excel (XLSX)[/bold]")
            menu.add_row("[white]   ├─ Formatted spreadsheet with styling[/white]")
            menu.add_row("[white]   └─ Best for detailed analysis[/white]")
            menu.add_row("")
            menu.add_row("[cyan][[/cyan]2[cyan]][/cyan] [bold]Export to CSV[/bold]")
            menu.add_row("[white]   ├─ Simple text-based format[/white]")
            menu.add_row("[white]   └─ Best for data portability[/white]")
            menu.add_row("")
            menu.add_row("[cyan][[/cyan]3[cyan]][/cyan] [bold]Back to Main Menu[/bold]")
            
            console.print(Panel(menu, title="[bold blue]Export Options[/bold blue]", border_style="cyan"))
            
            choice = Prompt.ask("Select export format", choices=["1", "2", "3"])
            
            if choice == "1":
                self.export_to_excel()
                input("\nPress Enter to continue...")
            elif choice == "2":
                self.export_to_csv()
                input("\nPress Enter to continue...")
            elif choice == "3":
                break

    def show_statistics(self):
        """Display statistics about scraped stories"""
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            
            cursor.execute('SELECT COUNT(*) FROM stories')
            total_stories = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(DISTINCT user_id) FROM stories')
            unique_users = cursor.fetchone()[0]
            
            cursor.execute('SELECT MAX(timestamp) FROM stories')
            last_story = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM stories WHERE date(timestamp) = date("now")')
            today_stories = cursor.fetchone()[0]
            
            conn.close()

            stats_table = Table(title="Stories Statistics", box=box.ROUNDED)
            stats_table.add_column("Metric", style="cyan")
            stats_table.add_column("Value", style="green")
            
            stats_table.add_row("Total Stories", str(total_stories))
            stats_table.add_row("Unique Users", str(unique_users))
            stats_table.add_row("Stories Today", str(today_stories))
            stats_table.add_row("Last Story Date", str(last_story))
            
            console.print(Panel(stats_table, border_style="cyan"))
            
        except Exception as e:
            console.print(f"[red]Error getting statistics: {str(e)}[/red]")
        
        input("\nPress Enter to continue...")

    def prompt_for_interval(self):
        """Prompt for checking interval"""
        return IntPrompt.ask("Enter the checking interval in seconds", default=60)

    def start_scraping(self, interval):
        """Start the scraping process"""
        console.print(f"[cyan]Starting scraper with {interval}-second interval[/cyan]")
        console.print("[yellow]Press Ctrl+C to stop and return to menu[/yellow]")
        
        try:
            if not os.path.exists('stories'):
                os.makedirs('stories')
            
            loop = asyncio.get_event_loop()
            
            loop.run_until_complete(self.scrape_stories())
            
            schedule.every(interval).seconds.do(
                lambda: loop.run_until_complete(self.scrape_stories())
            )
            
            while True:
                schedule.run_pending()
                time.sleep(1)
                
        except KeyboardInterrupt:
            schedule.clear()
            console.print("\n[yellow]Scraping stopped. Returning to menu...[/yellow]")
            time.sleep(2)
        except Exception as e:
            console.print(f"[red]Error during scraping: {str(e)}[/red]")
            time.sleep(2)

    def show_menu(self):
        """Display and handle the main menu"""
        try:
            loop = asyncio.get_event_loop()
            
            while True:
                console.clear()
                display_banner()
                
                menu = Table.grid(padding=1)
                menu.add_row("[cyan][[/cyan]1[cyan]][/cyan] [bold]Start Story Scraping[/bold]")
                menu.add_row("[cyan][[/cyan]2[cyan]][/cyan] [bold]Export Data[/bold]")
                menu.add_row("[cyan][[/cyan]3[cyan]][/cyan] [bold]View Statistics[/bold]")
                menu.add_row("[cyan][[/cyan]4[cyan]][/cyan] [bold]Change Interval[/bold]")
                menu.add_row("[cyan][[/cyan]5[cyan]][/cyan] [bold]Exit[/bold]")
                
                console.print(Panel(menu, title="[bold blue]Main Menu[/bold blue]", border_style="cyan"))
                
                choice = Prompt.ask("Select an option", choices=["1", "2", "3", "4", "5"])
                
                if choice == "1":
                    interval = self.prompt_for_interval()
                    self.start_scraping(interval)
                elif choice == "2":
                    self.export_data()
                elif choice == "3":
                    self.show_statistics()
                elif choice == "4":
                    interval = self.prompt_for_interval()
                    console.print(f"[green]Interval updated to {interval} seconds[/green]")
                    time.sleep(2)
                elif choice == "5":
                    if self.client and self.client.is_connected():
                        async def disconnect():
                            await self.client.disconnect()
                        loop.run_until_complete(disconnect())
                    console.print("[yellow]Goodbye![/yellow]")
                    break
                    
        except Exception as e:
            console.print(f"[red]Error in menu: {str(e)}[/red]")
def main():
    try:
        console.clear()
        display_banner()
        
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        except Exception:
            loop = asyncio.get_event_loop_policy().new_event_loop()
            asyncio.set_event_loop(loop)
            
        with asyncio.Runner() as runner:
            scraper = StoryScraper()
            runner.run(scraper.initialize_client())
            scraper.show_menu()
            
    except Exception as e:
        console.print(f"[red]Critical error: {str(e)}[/red]")
        input("Press Enter to exit...")
    finally:
        try:
            pending = asyncio.all_tasks(loop)
            loop.run_until_complete(asyncio.gather(*pending, return_exceptions=True))
        except Exception:
            pass
        finally:
            if loop and not loop.is_closed():
                loop.close()

if __name__ == "__main__":
    main()